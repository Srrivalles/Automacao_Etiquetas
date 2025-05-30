from tkinter import *
from tkinter import messagebox
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# Caminho do arquivo Excel
caminho_arquivo = os.path.join(os.getcwd(), "etiquetas_formatadas.xlsx")

# Variável para nome do projeto
nome_projeto = "PROJETO SEM NOME"

# Lista de projetos
projetos = ["Fracionado", "Entrega", "Coleta"]

# Função para salvar a lista de projetos
def salvar_projetos():
    with open("projetos.txt", "w") as arquivo:
        for projeto in projetos:
            arquivo.write(projeto + "\n")

# Função para carregar a lista de projetos
def carregar_projetos():
    if os.path.exists("projetos.txt"):
        with open("projetos.txt", "r") as arquivo:
            return [linha.strip() for linha in arquivo.readlines()]
    return []

# Carregar projetos ao iniciar
projetos = carregar_projetos()

# Função para criar etiquetas no Excel
def criar_etiquetas():
    try:
        cidade_estado = entrada_cidade_estado.get()
        fruta_cod1_cod2 = entrada_fruta_cod.get()
        quantidade_texto = entrada_quantidade.get()
        nota_fiscal = entrada_nota_fiscal.get()
        nome_responsavel = entrada_responsavel.get()
        obs = entrada_obs.get()

        # Validação dos campos obrigatórios
        import re
        if not re.match(r"^\d{2}/\d{2}$", quantidade_texto):
            raise ValueError("Formato inválido para quantidade. Use o formato 01/05.")

        if "/" not in quantidade_texto or not quantidade_texto.split("/")[1].isdigit():
            raise ValueError("Formato inválido para a quantidade. Deve ser no formato 01/05.")

        quantidade_total = int(quantidade_texto.split("/")[1])

        if os.path.exists(caminho_arquivo):
            workbook = load_workbook(caminho_arquivo)
        else:
            workbook = Workbook()
            workbook.active.title = "Etiquetas"

        sheet = workbook.active

        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        bold_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        regular_font = Font(name='Arial', size=12, color="FFFFFF")
        fill_color = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")


        linha_atual = sheet.max_row + 1 if sheet.max_row > 1 else 1

        for i in range(1, quantidade_total + 1):
            linha_inicial = linha_atual + (i - 1) * 14

            sheet.merge_cells(start_row=linha_inicial, start_column=1, end_row=linha_inicial + 1, end_column=3)
            cell_nome_projeto = sheet.cell(row=linha_inicial, column=1, value=nome_projeto)
            cell_nome_projeto.font = bold_font
            cell_nome_projeto.alignment = Alignment(horizontal='center', vertical='center')
            cell_nome_projeto.fill = fill_color

            for row in range(linha_inicial, linha_inicial + 2):
                for col in range(1, 4):
                    sheet.cell(row=row, column=col).border = border_style

            sheet.merge_cells(start_row=linha_inicial + 2, start_column=1, end_row=linha_inicial + 4, end_column=3)
            for row in range(linha_inicial + 2, linha_inicial + 5):
                for col in range(1, 4):
                    cell_empty = sheet.cell(row=row, column=col)
                    cell_empty.fill = fill_color
                    cell_empty.border = border_style

            labels = [
                ("Localidade:", cidade_estado),
                ("Nome do Equipamento:", fruta_cod1_cod2),
                ("Quantidade de Equipamentos:", f"{i:02d}/{quantidade_total:02d}"),
                ("Nota fiscal:", nota_fiscal),
                ("Responsável:", nome_responsavel),
                ("Obs:", obs)
            ]

            for j, (label_text, user_input) in enumerate(labels, start=linha_inicial + 5):
                label_cell = sheet.cell(row=j, column=1, value=label_text)
                label_cell.font = bold_font
                label_cell.fill = fill_color
                label_cell.alignment = Alignment(horizontal='left', vertical='center')

                user_input_cell = sheet.cell(row=j, column=2, value=user_input)
                user_input_cell.border = border_style

        sheet.column_dimensions['A'].width = 47.22
        sheet.column_dimensions['B'].width = 38.22
        sheet.column_dimensions['C'].width = 38.22

        workbook.save(caminho_arquivo)

        messagebox.showinfo("Sucesso", f"{quantidade_total} etiquetas criadas e salvas em: {caminho_arquivo}")
        print(f"[INFO] {quantidade_total} etiquetas geradas com sucesso.")

    except ValueError as ve:
        messagebox.showerror("Erro", f"Erro de validação: {ve}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Função para abrir a tela de gerenciamento de projetos
def gerenciar_projetos():
    janela_gerenciamento = Toplevel(root)
    janela_gerenciamento.title("Gerenciar Projetos")
    janela_gerenciamento.geometry("650x490")
    janela_gerenciamento.configure(bg="#f0f0f0")
    janela_gerenciamento.iconphoto(False, icone)  # Ícone aqui

    def adicionar_projeto():
        novo_projeto = entrada_novo_projeto.get()
        if novo_projeto:
            projetos.append(novo_projeto)
            listbox_projetos.insert(END, novo_projeto)
            entrada_novo_projeto.delete(0, END)
            salvar_projetos()
            messagebox.showinfo("Sucesso", f"Projeto '{novo_projeto}' adicionado com sucesso.")
        else:
            messagebox.showerror("Erro", "Por favor, insira um nome para o projeto.")

    def remover_projeto():
        selecionado = listbox_projetos.curselection()
        if selecionado:
            projeto_removido = listbox_projetos.get(selecionado)
            projetos.remove(projeto_removido)
            listbox_projetos.delete(selecionado)
            salvar_projetos()
            messagebox.showinfo("Sucesso", f"Projeto '{projeto_removido}' removido com sucesso.")
        else:
            messagebox.showerror("Erro", "Por favor, selecione um projeto para remover.")

    Label(janela_gerenciamento, text="Adicionar Novo Projeto:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=10)
    entrada_novo_projeto = Entry(janela_gerenciamento, font=("Segoe UI", 14))
    entrada_novo_projeto.pack(pady=5)

    Button(janela_gerenciamento, text="Adicionar Projeto", bg="#4CAF50", fg="white", command=adicionar_projeto, font=("Segoe UI", 14)).pack(pady=5)

    Label(janela_gerenciamento, text="Projetos Existentes:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=10)
    listbox_projetos = Listbox(janela_gerenciamento, font=("Segoe UI", 14))
    for projeto in projetos:
        listbox_projetos.insert(END, projeto)
    listbox_projetos.pack(pady=5)

    Button(janela_gerenciamento, text="Remover Projeto Selecionado", bg="#FF0000", fg="white", command=remover_projeto, font=("Segoe UI", 14)).pack(pady=10)

# Função para selecionar projeto
def selecionar_projeto():
    janela_selecao = Toplevel(root)
    janela_selecao.title("Selecionar Projeto")
    janela_selecao.geometry("450x400")
    janela_selecao.configure(bg="#f0f0f0")
    janela_selecao.iconphoto(False, icone)  # Ícone aqui

    def confirmar_selecao():
        global nome_projeto
        selecionado = listbox_projetos.curselection()
        if selecionado:
            nome_projeto = listbox_projetos.get(selecionado)
            messagebox.showinfo("Sucesso", f"Projeto selecionado: {nome_projeto}")
            janela_selecao.destroy()
        else:
            messagebox.showerror("Erro", "Por favor, selecione um projeto.")

    Label(janela_selecao, text="Selecione um Projeto:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=10)
    listbox_projetos = Listbox(janela_selecao, font=("Segoe UI", 14))
    for projeto in projetos:
        listbox_projetos.insert(END, projeto)
    listbox_projetos.pack(pady=10)

    Button(janela_selecao, text="Confirmar", bg="#4CAF50", fg="white", command=confirmar_selecao, font=("Segoe UI", 14)).pack(pady=10)

def alterar_nome_projeto():
    janela_escolha = Toplevel(root)
    janela_escolha.title("Selecione uma opção")
    janela_escolha.geometry("400x250")
    janela_escolha.configure(bg="#f0f0f0")
    janela_escolha.iconphoto(False, icone)  # Ícone aqui

    Label(janela_escolha, text="Escolha uma opção:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=20)

    Button(janela_escolha, text="Adicionar/Excluir Projeto", bg="#4CAF50", fg="white", command=gerenciar_projetos, font=("Segoe UI", 14)).pack(pady=10)
    Button(janela_escolha, text="Selecionar Projeto Existente", bg="#FFA500", fg="white", command=selecionar_projeto, font=("Segoe UI", 14)).pack(pady=10)

def limpar_campos():
    entrada_cidade_estado.delete(0, END)
    entrada_fruta_cod.delete(0, END)
    entrada_quantidade.delete(0, END)
    entrada_nota_fiscal.delete(0, END)
    entrada_responsavel.delete(0, END)
    entrada_obs.delete(0, END)
    messagebox.showinfo("Informação", "Campos limpos para adicionar um novo produto.")

# Interface principal
root = Tk()
root.title("Automação de Etiquetas")
root.geometry("880x592")
root.configure(bg="#f0f0f0")

# Carrega o ícone
icone = PhotoImage(file="etiqueta.png")
root.iconphoto(False, icone)

label_font = ('Segoe UI', 14)

# Frame para campos
frame_campos = Frame(root, bg="#f0f0f0")
frame_campos.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

# Campos
Label(frame_campos, text="Localidade:", bg="#f0f0f0", fg="black", font=label_font).grid(row=0, column=0, padx=10, pady=10, sticky=W)
entrada_cidade_estado = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_cidade_estado.grid(row=0, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Nome do Equipamento:", bg="#f0f0f0", fg="black", font=label_font).grid(row=1, column=0, padx=10, pady=10, sticky=W)
entrada_fruta_cod = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_fruta_cod.grid(row=1, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Quantidade (01/05):", bg="#f0f0f0", fg="black", font=label_font).grid(row=2, column=0, padx=10, pady=10, sticky=W)
entrada_quantidade = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_quantidade.grid(row=2, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Nota Fiscal:", bg="#f0f0f0", fg="black", font=label_font).grid(row=3, column=0, padx=10, pady=10, sticky=W)
entrada_nota_fiscal = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_nota_fiscal.grid(row=3, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Nome do Responsável:", bg="#f0f0f0", fg="black", font=label_font).grid(row=4, column=0, padx=10, pady=10, sticky=W)
entrada_responsavel = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_responsavel.grid(row=4, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Observações:", bg="#f0f0f0", fg="black", font=label_font).grid(row=6, column=0, padx=10, pady=10, sticky=W)
entrada_obs = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_obs.grid(row=6, column=1, padx=10, pady=10, sticky=W)

# Frame para botões
frame_botoes = Frame(root, bg="#f0f0f0")
frame_botoes.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

Button(frame_botoes, text="Criar Etiquetas", bg="#4CAF50", fg="white", command=criar_etiquetas, font=("Segoe UI", 14)).grid(row=0, column=0, padx=10, pady=10, sticky=W)
Button(frame_botoes, text="Alterar Nome do Projeto", bg="#FFA500", fg="white", command=alterar_nome_projeto, font=("Segoe UI", 14)).grid(row=0, column=1, padx=10, pady=10, sticky=W)
Button(frame_botoes, text="Limpar Campos", bg="#FF0000", fg="white", command=limpar_campos, font=("Segoe UI", 14)).grid(row=0, column=2, padx=10, pady=10, sticky=W)

# Iniciar
root.mainloop()
