from tkinter import *
from tkinter import messagebox
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# Caminho do arquivo Excel
caminho_arquivo = os.path.join(os.getcwd(), "etiquetas_formatadas.xlsx")

# Variável para nome do projeto
nome_projeto = "PROJETO SEM NOME"  # Nome inicial do projeto

# Lista de projetos
projetos = ["VENDER", "COMPRAR", "TRANSBORDO", "FILIAL", "FRANQUIA"]

# Função para salvar a lista de projetos 
def salvar_projetos():
    with open("projetos.txt", "w") as arquivo:
        for projeto in projetos:
            arquivo.write(projeto + "\n")

# Função para carregar a lista de projetos
def carregar_projetos():
    if os.path.exists("projetos.txt"):
        with open("projetos.txt", "r") as arquivo:
            return [linha.strip() for linha in arquivo.readlines()]
    return []

# Carregar projetos ao iniciar
projetos = carregar_projetos()

# Função para criar etiquetas no Excel
def criar_etiquetas():
    try:
        cidade_estado = entrada_cidade_estado.get()
        fruta_cod1_cod2 = entrada_fruta_cod.get()
        quantidade_texto = entrada_quantidade.get()
        nota_fiscal = entrada_nota_fiscal.get()
        nome_responsavel = entrada_responsavel.get()
        obs = entrada_obs.get()

        # Validação dos campos obrigatórios
        if not cidade_estado or not fruta_cod1_cod2 or not quantidade_texto or not nota_fiscal or not nome_responsavel:
            raise ValueError("Todos os campos obrigatórios devem ser preenchidos.")

        # Validação da entrada de quantidade no formato correto (ex: "01/05")
        if "/" not in quantidade_texto or not quantidade_texto.split("/")[1].isdigit():
            raise ValueError("Formato inválido para a quantidade. Deve ser no formato 01/05.")

        # Separar o número total de etiquetas
        quantidade_total = int(quantidade_texto.split("/")[1])

        # Criar ou carregar o arquivo Excel
        if os.path.exists(caminho_arquivo):
            workbook = load_workbook(caminho_arquivo)
        else:
            workbook = Workbook()
            workbook.active.title = "Etiquetas"

        sheet = workbook.active

        # Definir bordas expandidas
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Definir estilo de fonte
        bold_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        regular_font = Font(name='Arial', size=12, color="FFFFFF")

        # Cor de fundo
        fill_color = PatternFill(start_color="333399", end_color="333399", fill_type="solid")

        # Encontrar a próxima linha vazia
        linha_atual = sheet.max_row + 1 if sheet.max_row > 1 else 1

        # Loop para gerar etiquetas em linhas
        for i in range(1, quantidade_total + 1):
            linha_inicial = linha_atual + (i - 1) * 14  # Ajuste para incluir uma linha de espaço

            # Mesclar células para o nome do projeto com bordas e fundo
            sheet.merge_cells(start_row=linha_inicial, start_column=1, end_row=linha_inicial + 1, end_column=3)
            cell_nome_projeto = sheet.cell(row=linha_inicial, column=1, value=nome_projeto)
            cell_nome_projeto.font = bold_font
            cell_nome_projeto.alignment = Alignment(horizontal='center', vertical='center')
            cell_nome_projeto.fill = fill_color

            # Aplicar bordas nas células mescladas
            for row in range(linha_inicial, linha_inicial + 2):
                for col in range(1, 4):
                    sheet.cell(row=row, column=col).border = border_style

            # Mesclar três linhas abaixo, sem texto com bordas e fundo
            sheet.merge_cells(start_row=linha_inicial + 2, start_column=1, end_row=linha_inicial + 4, end_column=3)
            for row in range(linha_inicial + 2, linha_inicial + 5):
                for col in range(1, 4):
                    cell_empty = sheet.cell(row=row, column=col)
                    cell_empty.fill = fill_color
                    cell_empty.border = border_style

            # Adicionar a imagem dentro das células mescladas
            img = Image('TIM-LOGO.png')  # Substitua pelo caminho da sua imagem
            img.width = 70 
            img.height = 40 
            sheet.add_image(img, f'A{linha_inicial + 1}')  # Ajuste a célula conforme necessário

            # Adicionar informações do usuário
            labels = [
                ("Localidade:", cidade_estado),
                ("Área Responsável:", "EMPRESA:"),  # FIXO
                ("Nome do Equipamento:", fruta_cod1_cod2),
                ("Quantidade de Equipamentos:", f"{i:02d}/{quantidade_total:02d}"),
                ("Nota fiscal:", nota_fiscal),
                ("Responsável:", nome_responsavel),
                ("Demanda:", "Rollout"),  # Fixo
                ("Obs:", obs)
            ]

            # Loop para preencher as labels e as informações do usuário
            for j, (label_text, user_input) in enumerate(labels, start=linha_inicial + 5):
                # Definir as labels
                label_cell = sheet.cell(row=j, column=1, value=label_text)
                label_cell.font = bold_font
                label_cell.fill = fill_color
                label_cell.alignment = Alignment(horizontal='left', vertical='center')

                # Definir as entradas do usuário
                user_input_cell = sheet.cell(row=j, column=2, value=user_input)
                user_input_cell.border = border_style

        # Ajustar largura das colunas para melhor visualização
        sheet.column_dimensions['A'].width = 47.22
        sheet.column_dimensions['B'].width = 38.22
        sheet.column_dimensions['C'].width = 38.22

        # Salvar o arquivo Excel
        workbook.save(caminho_arquivo)

        # Exibir mensagem de sucesso com o caminho do arquivo
        messagebox.showinfo("Sucesso", f"{quantidade_total} etiquetas criadas e salvas em: {caminho_arquivo}")

    except ValueError as ve:
        messagebox.showerror("Erro", f"Erro de validação: {ve}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

    # Função para abrir a tela de gerenciamento de projetos
def gerenciar_projetos():
    # Criar uma nova janela para gerenciamento de projetos
    janela_gerenciamento = Toplevel(root)
    janela_gerenciamento.title("Gerenciar Projetos")
    janela_gerenciamento.geometry("650x490")
    janela_gerenciamento.configure(bg="#f0f0f0")

    # Função para adicionar projeto
    def adicionar_projeto():
        novo_projeto = entrada_novo_projeto.get()
        if novo_projeto:
            projetos.append(novo_projeto)
            listbox_projetos.insert(END, novo_projeto)
            entrada_novo_projeto.delete(0, END)
            salvar_projetos()  # Salvar a lista de projetos
            messagebox.showinfo("Sucesso", f"Projeto '{novo_projeto}' adicionado com sucesso.")
        else:
            messagebox.showerror("Erro", "Por favor, insira um nome para o projeto.")

    # Função para remover projeto
    def remover_projeto():
        selecionado = listbox_projetos.curselection()
        if selecionado:
            projeto_removido = listbox_projetos.get(selecionado)
            projetos.remove(projeto_removido)
            listbox_projetos.delete(selecionado)
            salvar_projetos()  # Salvar a lista de projetos
            messagebox.showinfo("Sucesso", f"Projeto '{projeto_removido}' removido com sucesso.")
        else:
            messagebox.showerror("Erro", "Por favor, selecione um projeto para remover.")

    # Campo de entrada para novo projeto
    Label(janela_gerenciamento, text="Adicionar Novo Projeto:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=10)
    entrada_novo_projeto = Entry(janela_gerenciamento, font=("Segoe UI", 14))
    entrada_novo_projeto.pack(pady=5)

    # Botão para adicionar projeto
    Button(janela_gerenciamento, text="Adicionar Projeto", bg="#4CAF50", fg="white", command=adicionar_projeto, font=("Segoe UI", 14)).pack(pady=5)

    # Listbox para exibir projetos existentes
    Label(janela_gerenciamento, text="Projetos Existentes:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=10)
    listbox_projetos = Listbox(janela_gerenciamento, font=("Segoe UI", 14))
    for projeto in projetos:
        listbox_projetos.insert(END, projeto)
    listbox_projetos.pack(pady=5)

    # Botão para remover projeto
    Button(janela_gerenciamento, text="Remover Projeto Selecionado", bg="#FF0000", fg="white", command=remover_projeto, font=("Segoe UI", 14)).pack(pady=10)

    # Função para selecionar um projeto existente
def selecionar_projeto():
    # Criar uma nova janela para seleção de projetos
    janela_selecao = Toplevel(root)
    janela_selecao.title("Selecionar Projeto")
    janela_selecao.geometry("450x400")
    janela_selecao.configure(bg="#f0f0f0")

    # Função para confirmar a seleção
    def confirmar_selecao():
        global nome_projeto
        selecionado = listbox_projetos.curselection()
        if selecionado:
            nome_projeto = listbox_projetos.get(selecionado)
            messagebox.showinfo("Sucesso", f"Projeto selecionado: {nome_projeto}")
            janela_selecao.destroy()
        else:
            messagebox.showerror("Erro", "Por favor, selecione um projeto.")

    # Listbox para exibir projetos existentes
    Label(janela_selecao, text="Selecione um Projeto:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=10)
    listbox_projetos = Listbox(janela_selecao, font=("Segoe UI", 14))
    for projeto in projetos:
        listbox_projetos.insert(END, projeto)
    listbox_projetos.pack(pady=10)

    # Botão para confirmar seleção
    Button(janela_selecao, text="Confirmar", bg="#4CAF50", fg="white", command=confirmar_selecao, font=("Segoe UI", 14)).pack(pady=10)

# Função para abrir a primeira tela de escolha
def alterar_nome_projeto():
    # Criar uma nova janela para escolha
    janela_escolha = Toplevel(root)
    janela_escolha.title("Selecione uma opção")
    janela_escolha.geometry("400x250")
    janela_escolha.configure(bg="#f0f0f0")

    # Mensagem de instrução
    Label(janela_escolha, text="Escolha uma opção:", bg="#f0f0f0", fg="black", font=("Segoe UI", 14)).pack(pady=20)

    # Botão para gerenciar projetos (adicionar/excluir)
    Button(janela_escolha, text="Adicionar/Excluir Projeto", bg="#4CAF50", fg="white", command=gerenciar_projetos, font=("Segoe UI", 14)).pack(pady=10)

    # Botão para selecionar projeto existente
    Button(janela_escolha, text="Selecionar Projeto Existente", bg="#FFA500", fg="white", command=selecionar_projeto, font=("Segoe UI", 14)).pack(pady=10)

# Função para limpar os campos
def limpar_campos():
    entrada_cidade_estado.delete(0, END)
    entrada_fruta_cod.delete(0, END)
    entrada_quantidade.delete(0, END)
    entrada_nota_fiscal.delete(0, END)
    entrada_responsavel.delete(0, END)
    entrada_obs.delete(0, END)
    messagebox.showinfo("Informação", "Campos limpos para adicionar um novo produto.")

# Interface gráfica principal
root = Tk()
root.title("Automação de Etiquetas")
root.geometry("880x592")
root.configure(bg="#f0f0f0")  # Fundo cinza claro

# Estilo para as labels
label_font = ('Segoe UI', 14)

# Frame para os campos de entrada
frame_campos = Frame(root, bg="#f0f0f0")
frame_campos.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

# Campos de entrada
Label(frame_campos, text="Localidade:", bg="#f0f0f0", fg="black", font=label_font).grid(row=0, column=0, padx=10, pady=10, sticky=W)
entrada_cidade_estado = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_cidade_estado.grid(row=0, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Nome do Equipamento:", bg="#f0f0f0", fg="black", font=label_font).grid(row=1, column=0, padx=10, pady=10, sticky=W)
entrada_fruta_cod = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_fruta_cod.grid(row=1, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Quantidade (f-(01/05)):", bg="#f0f0f0", fg="black", font=label_font).grid(row=2, column=0, padx=10, pady=10, sticky=W)
entrada_quantidade = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_quantidade.grid(row=2, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Nota Fiscal:", bg="#f0f0f0", fg="black", font=label_font).grid(row=3, column=0, padx=10, pady=10, sticky=W)
entrada_nota_fiscal = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_nota_fiscal.grid(row=3, column=1, padx=10, pady=10, sticky=W)

Label(frame_campos, text="Nome do Responsável:", bg="#f0f0f0", fg="black", font=label_font).grid(row=4, column=0, padx=10, pady=10, sticky=W)
entrada_responsavel = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_responsavel.grid(row=4, column=1, padx=10, pady=10, sticky=W)


Label(frame_campos, text="Observações:", bg="#f0f0f0", fg="black", font=label_font).grid(row=6, column=0, padx=10, pady=10, sticky=W)
entrada_obs = Entry(frame_campos, bg="white", fg="black", width=50, font=label_font)
entrada_obs.grid(row=6, column=1, padx=10, pady=10, sticky=W)

# Frame para os botões
frame_botoes = Frame(root, bg="#f0f0f0")
frame_botoes.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

# Botões para ações
Button(frame_botoes, text="Criar Etiquetas", bg="#4CAF50", fg="white", command=criar_etiquetas, font=("Segoe UI", 14)).grid(row=0, column=0, padx=10, pady=10, sticky=W)
Button(frame_botoes, text="Alterar Nome do Projeto", bg="#FFA500", fg="white", command=alterar_nome_projeto, font=("Segoe UI", 14)).grid(row=0, column=1, padx=10, pady=10, sticky=W)
Button(frame_botoes, text="Limpar Campos", bg="#FF0000", fg="white", command=limpar_campos, font=("Segoe UI", 14)).grid(row=0, column=2, padx=10, pady=10, sticky=W)

# Iniciar a interface gráfica
root.mainloop()
